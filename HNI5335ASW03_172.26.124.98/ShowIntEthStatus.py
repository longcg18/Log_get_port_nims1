import textfsm as tf
import pandas as pd
import openpyxl
import xlsxwriter

def interface_name(name):
    ats = str(name)
    return ats.replace("FE", "FastEthernet").replace("G-", "Gigabit-")

def read_data(fileName):
    f = open(fileName, "r")
    start_marker = "Interface  Link/Protocol  Speed   Duplex  Vlan   Type"
    end_marker = "@BLOCK--"

    file_content = f.read()

    start_index = file_content.find(start_marker) + len(start_marker)
    end_index = file_content.find(end_marker, start_index)

    relevant_data = file_content[start_index:end_index].strip()
    return relevant_data

if (__name__ == '__main__'):
    with open("ShowIntEthStatus.template") as tpl:
        fsm = tf.TextFSM(tpl)
    
    data = read_data("HNI5335ASW03_172.26.124.98_DCS-3950.txt")
    results = fsm.ParseText(data)
    fileName = "HNI5335ASW03_172.26.124.98_DCS-3950.txt"
    deviceInfos = fileName.split("_")
    deviceName = deviceInfos[0]
    deviceIP = deviceInfos[1]
    deviceModel = deviceInfos[2]

    workbook = openpyxl.load_workbook("..\DataCollection.xlsx")
    sheetName = str(deviceName)
    if (sheetName in workbook.sheetnames) == True:
        workbook.remove(workbook[sheetName]) 
    worksheet = workbook.create_sheet(sheetName)
    title = ["Interface", "Link/Protocol State", "Speed", "Duplex", "Vlan", "AliasName"]

    for col, val in enumerate(title, start=1):
        worksheet.cell(row=1, column=col).value = val

    row_num=2
    for result in results:
        
        line = []
        line.append(str(interface_name(result[6]) + " " + result[0]))        
        line.append(str(result[1] + "/" + result[2]))        
        line.append(str(result[3]))        
        line.append(str(result[4]))
        line.append(str(result[5]))
        line.append(str(result[7]))
        
        for col_num, res in enumerate(line, start=1):
            worksheet.cell(row_num, col_num).value=res
        row_num = row_num + 1
    
    workbook.save("..\DataCollection.xlsx")
    workbook.close()

